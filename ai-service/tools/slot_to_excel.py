from openpyxl import load_workbook


def generate_parking_slots(
        template_excel_path: str,
        output_excel_path: str,
        area_id: str,
        camera_id: str,
        slot_count: int,
        start_slot_code: int = 1,
        default_status: str = "FREE",
        slot_code_prefix: str = ""
):
    """
    向车位导入模板中批量写入车位信息。

    :param template_excel_path: 模板 Excel 路径（含标题）
    :param output_excel_path: 输出 Excel 路径
    :param area_id: 停车区域ID（如 roi_2）
    :param camera_id: 绑定摄像头（数据库中的 camera_id）
    :param slot_count: 生成多少个车位
    :param start_slot_code: 起始车位编号（默认 1）
    :param default_status: 车位状态（默认 FREE）
    :param slot_code_prefix: 车位编号前缀（例如 'A-', 'B-', 'T1-'），默认为空字符串
    """

    wb = load_workbook(template_excel_path)
    ws = wb.active

    start_row = 2

    for i in range(slot_count):
        row = start_row + i

        slot_code_num = start_slot_code + i
        full_slot_code = f"{slot_code_prefix}{slot_code_num}"

        ws.cell(row=row, column=1, value=area_id)
        ws.cell(row=row, column=2, value=full_slot_code)
        ws.cell(row=row, column=3, value=default_status)
        ws.cell(row=row, column=4, value=str(camera_id))

    wb.save(output_excel_path)
    print(f"成功生成 {slot_count} 个车位，文件已保存：{output_excel_path}")


if __name__ == "__main__":
    TEMPLATE_EXCEL = r"E:/下载/车位信息_1768550124054.xlsx"
    OUTPUT_EXCEL = r"E:/下载/车位信息_roi_3.xlsx"

    AREA_ID = "3"
    CAMERA_ID = "2011736637179957250"
    SLOT_COUNT = 36
    START_SLOT_CODE = 1
    DEFAULT_STATUS = "空闲"
    SLOT_CODE_PREFIX = "停车场C-"
    generate_parking_slots(
        template_excel_path=TEMPLATE_EXCEL,
        output_excel_path=OUTPUT_EXCEL,
        area_id=AREA_ID,
        camera_id=CAMERA_ID,
        slot_count=SLOT_COUNT,
        start_slot_code=START_SLOT_CODE,
        default_status=DEFAULT_STATUS,
        slot_code_prefix=SLOT_CODE_PREFIX
    )
