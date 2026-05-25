import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


def export_roi_to_excel(
    roi_json_path: Path,
    excel_template_path: Path,
    camera_name: str,
    slot_code_prefix: str,
) -> None:
    with roi_json_path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    image_width, image_height = data["image_size"]
    slots = data["slots"]

    workbook = load_workbook(excel_template_path)
    worksheet = workbook.active

    for idx, slot in enumerate(slots, start=2):
        slot_code = f"{slot_code_prefix}{slot['slot_id']}"
        roi_polygon = json.dumps(slot["polygon"], ensure_ascii=False)

        worksheet.cell(row=idx, column=1, value=slot_code)
        worksheet.cell(row=idx, column=2, value=camera_name)
        worksheet.cell(row=idx, column=3, value=image_width)
        worksheet.cell(row=idx, column=4, value=image_height)
        worksheet.cell(row=idx, column=5, value=roi_polygon)

    workbook.save(excel_template_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="将 ROI JSON 写入车位导入 Excel 模板")
    parser.add_argument("--roi-json", required=True, type=Path, help="ROI JSON 文件路径")
    parser.add_argument("--excel", required=True, type=Path, help="Excel 模板文件路径")
    parser.add_argument("--camera-name", required=True, help="摄像头名称")
    parser.add_argument("--slot-prefix", default="", help="车位编号前缀")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    export_roi_to_excel(
        roi_json_path=args.roi_json,
        excel_template_path=args.excel,
        camera_name=args.camera_name,
        slot_code_prefix=args.slot_prefix,
    )
    print("ROI 数据已写入模板 Excel")


if __name__ == "__main__":
    main()
