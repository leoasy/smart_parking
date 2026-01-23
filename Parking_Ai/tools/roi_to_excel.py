import json
from openpyxl import load_workbook

# ========== 你需要改的 ==========
ROI_JSON_PATH = "D:/PycharmProjects/Parking_Ai/data/roi/roi_3_camera_1.json"
EXCEL_TEMPLATE_PATH = "E:/下载/车位ROI标注_1768637462643.xlsx"
CAMERA_NAME = "停车场C-广角"
SLOT_CODE_PREFIX = "停车场C-"   # 如果是 A-1 这种，就写 "A-"
# ===============================


def main():
    # 1️⃣ 读取 ROI JSON
    with open(ROI_JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    image_width, image_height = data["image_size"]
    slots = data["slots"]

    # 2️⃣ 打开已有模板
    wb = load_workbook(EXCEL_TEMPLATE_PATH)
    ws = wb.active

    start_row = 2  # 从第 2 行开始写

    for idx, slot in enumerate(slots):
        row = start_row + idx

        slot_code = f"{SLOT_CODE_PREFIX}{slot['slot_id']}"
        roi_polygon = json.dumps(slot["polygon"], ensure_ascii=False)

        # 按模板列顺序写（不要乱）
        ws.cell(row=row, column=1, value=slot_code)      # 车位编号
        ws.cell(row=row, column=2, value=CAMERA_NAME)    # 摄像头名称
        ws.cell(row=row, column=3, value=image_width)    # 图片宽度
        ws.cell(row=row, column=4, value=image_height)   # 图片高度
        ws.cell(row=row, column=5, value=roi_polygon)    # ROI多边形坐标

    wb.save(EXCEL_TEMPLATE_PATH)
    print("✅ ROI 数据已写入模板 Excel")


if __name__ == "__main__":
    main()
